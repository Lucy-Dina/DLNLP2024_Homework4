import numpy as np
import random
import matplotlib.pyplot as plt
import matplotlib
import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader
import pickle
import pdb
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

class CorpusDataset(Dataset):
    def __init__(self, source_data, target_data, source_word_2_idx, target_word_2_idx):
        self.source_data = source_data
        self.target_data = target_data
        self.source_word_2_idx = source_word_2_idx
        self.target_word_2_idx = target_word_2_idx

    def __getitem__(self, index):
        # create one-hot index for every word
        src = self.source_data[index]
        tgt = self.target_data[index]

        src_index = [self.source_word_2_idx[i] for i in src]
        tgt_index = [self.target_word_2_idx[i] for i in tgt]

        return src_index, tgt_index

    def batch_data_alignment(self, batch_datas):
        global device
        src_index, tgt_index = [], []
        src_len, tgt_len = [], []

        for src, tgt in batch_datas:
            src_index.append(src)
            tgt_index.append(tgt)
            src_len.append(len(src))
            tgt_len.append(len(tgt))

        max_src_len = max(src_len)
        max_tgt_len = max(tgt_len)
        src_index = [[self.source_word_2_idx["<BOS>"]] + tmp_src_index + [self.source_word_2_idx["<EOS>"]] +
                     [self.source_word_2_idx["<PAD>"]] * (max_src_len - len(tmp_src_index)) for tmp_src_index in
                     src_index] #补齐标识符“<PAD>”，使文本序列的长度与该批次数据中最长文本序列的长度一致
        tgt_index = [[self.target_word_2_idx["<BOS>"]] + tmp_src_index + [self.target_word_2_idx["<EOS>"]] +
                     [self.target_word_2_idx["<PAD>"]] * (max_tgt_len - len(tmp_src_index)) for tmp_src_index in
                     tgt_index]
        src_index = torch.tensor(src_index, device=device)
        tgt_index = torch.tensor(tgt_index, device=device)

        return src_index, tgt_index

    def __len__(self):
        assert len(self.source_data) == len(self.target_data)
        return len(self.target_data)


class Encoder(nn.Module):
    def __init__(self, dim_encoder_embbeding, dim_encoder_hidden, source_corpus_len):
        super().__init__()
        self.embedding = nn.Embedding(source_corpus_len, dim_encoder_embbeding)
        self.lstm = nn.LSTM(dim_encoder_embbeding, dim_encoder_hidden, batch_first=True)    #LSTM

    def forward(self, src_index):
        en_embedding = self.embedding(src_index)
        _, encoder_hidden = self.lstm(en_embedding)

        return encoder_hidden


class Decoder(nn.Module):
    def __init__(self, dim_decoder_embedding, dim_decoder_hidden, target_corpus_len):
        super().__init__()
        self.embedding = nn.Embedding(target_corpus_len, dim_decoder_embedding)
        self.lstm = nn.LSTM(dim_decoder_embedding, dim_decoder_hidden, batch_first=True)

    def forward(self, decoder_input, hidden):
        embedding = self.embedding(decoder_input)
        decoder_output, decoder_hidden = self.lstm(embedding, hidden)

        return decoder_output, decoder_hidden


class Seq2Seq(nn.Module):
    def __init__(self, dim_encoder_embbeding, dim_encoder_hidden, source_corpus_len,
                 dim_decoder_embedding, dim_decoder_hidden, target_corpus_len):
        super().__init__()
        self.encoder = Encoder(dim_encoder_embbeding, dim_encoder_hidden, source_corpus_len)
        self.decoder = Decoder(dim_decoder_embedding, dim_decoder_hidden, target_corpus_len)
        self.classifier = nn.Linear(dim_decoder_hidden, target_corpus_len)

        self.ce_loss = nn.CrossEntropyLoss()

    def forward(self, src_index, tgt_index):
        decoder_input = tgt_index[:, :-1]
        label = tgt_index[:, 1:]

        encoder_hidden = self.encoder(src_index)
        decoder_output, _ = self.decoder(decoder_input, encoder_hidden)

        pre = self.classifier(decoder_output)
        loss = self.ce_loss(pre.reshape(-1, pre.shape[-1]), label.reshape(-1))

        return loss


def generate_sentence(sentence):
    global source_word_2_idx, model, device, target_word_2_idx, target_idx_2_word

    src_index = torch.tensor([[source_word_2_idx[i] for i in sentence]], device=device)

    result = []
    encoder_hidden = model.encoder(src_index)
    decoder_input = torch.tensor([[target_word_2_idx["<BOS>"]]], device=device)

    decoder_hidden = encoder_hidden
    while True:
        decoder_output, decoder_hidden = model.decoder(decoder_input, decoder_hidden)
        pre = model.classifier(decoder_output)

        w_index = int(torch.argmax(pre, dim=-1))
        word = target_idx_2_word[w_index]

        if word == "<EOS>" or len(result) > 40:
            break

        result.append(word)
        decoder_input = torch.tensor([[w_index]], device=device)

    return "".join(result)


if __name__ == '__main__':

    # ====== Initialize ======
    device = "cuda:0" if torch.cuda.is_available() else "cpu"
    batch_size = 2
    num_corpus = 300    # length of train set
    num_test_corpus = 10    # # length of test set
    txt_file_path = "E:/Deep_NLP/Homework4/天龙八部.txt"
    # txt_file_path = "D:/Desktop/zzzzzz/2024_spring/Deep_NLP/Homework4/天龙八部.txt"
    num_epochs = 50
    lr = 0.001
    dim_encoder_embbeding = 150
    dim_encoder_hidden = 100
    dim_decoder_embedding = 150
    dim_decoder_hidden = 100

    # ====== Data Preprocess ======
    char_to_be_replaced = "\n 0123456789qwertyuiopasdfghjklzxcvbnm[]{};':\",./<>?ａｎｔｉ－ｃｌｉｍａｘ＋．／０１２３４５６７８９＜＝＞＠Ａ" \
                          "ＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＶＷＸＹＺ［＼］ｂｄｅｆｇｈｊｋｏｐｒｓ" \
                          "ｕｖｗｙｚ￣\u3000\x1a"

    source_target_corpus_ori = []
    # filter
    with open(txt_file_path, "r", encoding="gbk", errors="ignore") as tmp_file:
        tmp_file_context = tmp_file.read()
        for tmp_char in char_to_be_replaced:
            tmp_file_context = tmp_file_context.replace(tmp_char, "")
        tmp_file_context = tmp_file_context.replace("本书来自免费小说下载站更多更新免费电子书请关注", "")
        # sample
        tmp_file_sentences = tmp_file_context.split("。")
        for tmp_idx, tmp_sentence in enumerate(tmp_file_sentences): # 根据设定好的采样规则筛选出所有符合要求的句子
            if ("段" in tmp_sentence) and (10 <= len(tmp_sentence) <= 40) and (
                    10 <= len(tmp_file_sentences[tmp_idx + 1]) <= 40):
                # 依次存入被采样的句子及其下一句话，即label
                source_target_corpus_ori.append((tmp_file_sentences[tmp_idx], tmp_file_sentences[tmp_idx + 1]))

    # ====== Build Train & Test Set ======
    sample_indexes = random.sample(list(range(len(source_target_corpus_ori))), num_corpus)
    source_corpus, target_corpus = [], []
    for idx in sample_indexes:
        source_corpus.append(source_target_corpus_ori[idx][0])  # train set
        target_corpus.append(source_target_corpus_ori[idx][1])  # train label

    test_corpus = []
    for idx in range(len(source_target_corpus_ori)):
        if idx not in sample_indexes:   # 保证训练集和测试集不同
            test_corpus.append((source_target_corpus_ori[idx][0], source_target_corpus_ori[idx][1]))
    test_corpus = random.sample(test_corpus, num_test_corpus)
    test_source_corpus, test_target_corpus = [], []
    for tmp_src, tmp_tgt in test_corpus:
        test_source_corpus.append(tmp_src)  # test set
        test_target_corpus.append(tmp_tgt)  # test label

    # ====== Generate One-Hot Dict ======
    idx_cnt = 0
    word_2_idx_dict = dict()
    idx_2_word_list = list()
    for tmp_corpus in [source_corpus, target_corpus, test_source_corpus, test_target_corpus]:
        for tmp_sentence in tmp_corpus:
            for tmp_word in tmp_sentence:   # 以字符为单位
                if tmp_word not in word_2_idx_dict.keys():
                    word_2_idx_dict[tmp_word] = idx_cnt
                    idx_2_word_list.append(tmp_word)
                    idx_cnt += 1

    one_hot_dict_len = len(word_2_idx_dict)
    # 再加上"<PAD>", "<BOS>", "<EOS>"
    word_2_idx_dict.update({"<PAD>": one_hot_dict_len, "<BOS>": one_hot_dict_len + 1, "<EOS>": one_hot_dict_len + 2})
    idx_2_word_list += ["<PAD>", "<BOS>", "<EOS>"]
    one_hot_dict_len += 3

    source_word_2_idx, target_word_2_idx = word_2_idx_dict, word_2_idx_dict
    source_idx_2_word, target_idx_2_word = idx_2_word_list, idx_2_word_list
    source_corpus_len, target_corpus_len = one_hot_dict_len, one_hot_dict_len

    # ====== Train Model ======
    # load train set
    dataset = CorpusDataset(source_corpus, target_corpus, source_word_2_idx, target_word_2_idx)
    dataloader = DataLoader(dataset, batch_size, shuffle=False, collate_fn=dataset.batch_data_alignment)
    # train
    model = Seq2Seq(dim_encoder_embbeding,
                    dim_encoder_hidden,
                    source_corpus_len,
                    dim_decoder_embedding,
                    dim_decoder_hidden,
                    target_corpus_len)
    model = model.to(device)    # 保证设备一致
    # optimize
    optimizer = torch.optim.Adam(model.parameters(), lr=lr)
    losses = []
    # iteration
    content = "epoch"
    ws.cell(row=1, column=1, value=content)
    content = "loss"
    ws.cell(row=1, column=2, value=content)
    for epoch in range(num_epochs):
        for step, (src_index, tgt_index) in enumerate(dataloader):
            loss = model(src_index, tgt_index)
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
        # 将损失值转为CPU上的Python标量
        losses.append(loss.item())  # 使用.item()从单个张量中提取数字
        # losses.append(loss)
        print("epoch: {}, training loss: {:.5f}".format(epoch + 1, loss))
        content = "{}".format(epoch + 1)
        ws.cell(row=epoch + 2, column=1, value=content)
        content = "{:.5f}".format(loss)
        ws.cell(row=epoch + 2, column=2, value=content)

    # ### visualize training loss
    # plt.figure()
    # plt.plot(np.array([i + 1 for i in range(num_epochs)]), losses, "b-", label="Training Loss")
    #
    # plt.legend()
    # plt.xlabel("Epoch")
    # plt.ylabel("Training Loss")
    # plt.title("Training Loss of Seq2Seq")
    #
    # plt.savefig("./training_loss.png")

    # ====== Test Model ======
    flag = num_epochs + 2
    for idx, (tmp_src_sentence, tmp_gt_sentence) in enumerate(test_corpus):
        tmp_generated_sentence = generate_sentence(tmp_src_sentence)
        print("----------------Result {}----------------".format(idx + 1))
        print("Source sentence: {}".format(tmp_src_sentence))
        print("True target sentence: {}".format(tmp_gt_sentence))
        print("Generated target sentence: {}".format(tmp_generated_sentence))
        content = "Result {}".format(idx + 1)
        ws.cell(row=flag, column=1, value=content)

        content = "Source sentence"
        ws.cell(row=flag + 1, column=1, value=content)
        content = "{}".format(tmp_src_sentence)
        ws.cell(row=flag + 1, column=2, value=content)

        content = "True target sentence"
        ws.cell(row=flag + 2, column=1, value=content)
        content = "{}".format(tmp_gt_sentence)
        ws.cell(row=flag + 2, column=2, value=content)

        content = "Generated target sentence"
        ws.cell(row=flag + 3, column=1, value=content)
        content = "{}".format(tmp_generated_sentence)
        ws.cell(row=flag + 3, column=2, value=content)

        flag = flag + 4

wb.save("outcome_Seq2Seq.xlsx")

